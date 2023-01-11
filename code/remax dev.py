from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import element_to_be_clickable
import openpyxl
import time
from datetime import date
import re
from selenium.webdriver.common.keys import Keys 

from email.message import EmailMessage
import ssl
import smtplib
import os

import sqlite3

# Connect to the database
conn = sqlite3.connect('remax.db')
# Create a cursor to execute SQL statements
cursor = conn.cursor()

# Create the table
cursor.execute('''CREATE TABLE IF NOT EXISTS remax (date text, total_flats integer, flats_s integer, flats_m integer, flats_l integer, flats_xl integer, houses integer)''')


#TODO
    

file_path = fr"c:\Users\halik\OneDrive\Plocha\remax web scraper\remax.xlsx"
url_byty = "https://www.sreality.cz/hledani/byty"
url_domy = "https://www.sreality.cz/hledani/domy"
urls = ["https://www.sreality.cz/hledani/byty", "https://www.sreality.cz/hledani/domy"]
chars = fr"abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ!$%&'()*+,-./:;<=>?@[\]^_`~ \t\n\r\x0b\x0c"
checks = []
velikosti_od = [0, 42, 67, 87]
velikosti_do = [41, 66, 86, 9999]
test_mode = False
if test_mode:
    sleep = 5
else:
    sleep = 1

def chrome(urls):
    dela_byty = True
    byty = [0, 0, 0, 0]
    for url in urls:
        # Set up the webdriver
        driver = webdriver.Chrome()
        # create a new Chrome browser instance in headless mode
        options = webdriver.ChromeOptions()
        if not test_mode:
            options.add_argument("--headless")
        driver = webdriver.Chrome(chrome_options=options)
        # Navigate to the URL
        driver.get(url)
        # Wait for the page to load
        time.sleep(sleep)
        # Find the button with the specified CSS class and click it
        button = driver.find_element(By.CSS_SELECTOR, ".praha .tspan")
        button.click()
        if dela_byty and len(checks) != 0:
            for check in checks:
                #find the checkbox
                checkbox = driver.find_element(By.LINK_TEXT, check)
                print(checkbox)
                # click on the 3px right and 10px down from the text
                action = webdriver.common.action_chains.ActionChains(driver)
                action.move_to_element_with_offset(checkbox, 3, -10)
                action.click()
                action.perform()
                # Wait for the page to load
                time.sleep(sleep)
        else:
            pass
        if dela_byty:
            od = driver.find_element(By.XPATH, '//*[@id="page-layout"]/div[2]/div[3]/div[3]/div/div/div[1]/form/div[1]/div[20]/div[2]/div/div/input[1]')
            do = driver.find_element(By.XPATH, '//*[@id="page-layout"]/div[2]/div[3]/div[3]/div/div/div[1]/form/div[1]/div[20]/div[2]/div/div/input[2]')
            for index in range(0, len(velikosti_od)):
                #write into the size form
                od.send_keys(Keys.CONTROL, "a")
                od.send_keys(velikosti_od[index])
                do.send_keys(Keys.CONTROL, "a")
                do.send_keys(velikosti_do[index])
                
                action = webdriver.common.action_chains.ActionChains(driver)
                action.move_to_element_with_offset(od, 30, -30)
                action.click()
                action.perform()
                # Wait for the page to load
                time.sleep(sleep)
                
                # Get all the text on the page
                page_text = driver.find_element(By.TAG_NAME, "body").text
                page_index = page_text.find("inzerátů")
                byty[index] = page_text[page_index - 7 : page_index]
                byty[index] = re.sub(r'[^0-9]', '', byty[index])
                print(f"byty{byty}")
                byty[index] = int(byty[index].replace(" ", ""))
                
        else:
            time.sleep(sleep)
            # Get all the text on the page
            page_text = driver.find_element(By.TAG_NAME, "body").text
            page_index = page_text.find("inzerátů")
            page_index = page_text.find("inzerátů")
            domy = page_text[page_index - 7 : page_index]
            print(f"domy str {domy}")
            domy = re.sub(r'[^0-9]', '', domy)
            print(f"domy{domy}")
            domy = int(domy.replace(" ", ""))
            
        dela_byty = False
        
        
    date_string = str(date)
    total_flats = sum(byty)
    flats_s = byty[0]
    flats_m = byty[1]
    flats_l = byty[2]
    flats_xl = byty[3]
    houses = domy
    # Insert the data into the table
    cursor.execute("INSERT INTO remax VALUES (?, ?, ?, ?, ?, ?, ?)", (date_string, total_flats, flats_s, flats_m, flats_l, flats_xl, houses))
    # Save the changes to the database
    conn.commit()
    # Close the connection
    conn.close()
    return byty,domy
    driver.quit()

def excel(byty, domy, file_path):
    byty_celkem = sum(byty)
    # Open the workbook and specify the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[2]
    # Specify the cell to read the value from
    cell = sheet['A1']
    # Read the value of the cell and ad one to it
    iteration = cell.value
    iteration += 1
    cell.value = iteration
    byty_celkem_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+2)
    byty_s_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+3)
    byty_m_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+4)
    byty_l_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+5)
    byty_xl_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+6)
    domy_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+7)
    date_cell = sheet.cell(cell.row+iteration, cell.col_idx+1)
    
    
    # Write the value num to the new cell
    byty_celkem_cell.value = byty_celkem
    byty_s_cell.value = byty[0]
    byty_m_cell.value = byty[1]
    byty_l_cell.value = byty[2]
    byty_xl_cell.value = byty[3]
    domy_cell.value = domy
    date_cell.value = date.today()
    # Save the changes to the workbook
    wb.save(file_path)
    print(f"file saved to {file_path}")
    return byty_celkem, domy, iteration
    
def statistiky(iteration):
    # Open the workbook and specify the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[2]
    cell = sheet['A1']
    #byty dnes
    byty_celkem_dnes = sheet.cell(row=cell.row+iteration, column=cell.col_idx+2).value
    domy_dnes = sheet.cell(row=cell.row+iteration, column=cell.col_idx+7).value
    #byty vcera
    byty_celkem_vcera_cell = sheet.cell(row=cell.row+iteration-1, column=cell.col_idx+2)
    byty_celkem_vcera = byty_celkem_vcera_cell.value
    byty_celkem_vcera_graph = sheet['G2']
    byty_celkem_vcera_graph.value = byty_celkem_vcera
    #datum posledni mesic
    datum_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+1).value for i in range(0,360)]
    for i in range(0,360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+11).value = datum_posledni_mesic[i]
    #byty posledni mesic
    byty_celkem_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+2).value for i in range(0,360)]
    for i in range(0,360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+12).value = byty_celkem_posledni_mesic[i]
    byty_s_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+3).value for i in range(0,360)]
    for i in range(0,360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+13).value = byty_s_posledni_mesic[i]
    byty_m_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+4).value for i in range(0,360)]
    for i in range(0,360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+14).value = byty_m_posledni_mesic[i]
        byty_l_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+5).value for i in range(0,360)]
    for i in range(0,360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+15).value = byty_l_posledni_mesic[i]        
        byty_xl_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+6).value for i in range(0,360)]
    for i in range(0,360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+16).value = byty_xl_posledni_mesic[i]
    #domy posledni mesic
    domy_posledni_mesic = [sheet.cell(row=cell.row+iteration-i,column=cell.col_idx+7).value for i in range(0,360)]
    for i in range(0, 360):
        sheet.cell(row=cell.row+i+1,column=cell.col_idx+17).value = domy_posledni_mesic[i]
    # Save the changes to the workbook
    wb.save(file_path)
    
    byty_pred_tydnem = int(sheet.cell(row=cell.row+8, column=cell.col_idx+12).value)
    byty_pred_mesicem = int(sheet.cell(row=cell.row+31, column=cell.col_idx+12).value)
    byty_tydenni_zmena = (byty_celkem_dnes - byty_pred_tydnem)*100/byty_celkem_dnes
    byty_mesicni_zmena = (byty_celkem_dnes - byty_pred_mesicem)*100/byty_celkem_dnes
    email_text = f"""
    dnes je v Praze v nabídce celkem {byty_celkem_dnes} bytů, a {domy_dnes} domů.
    Byty = týdenní změna {str(byty_tydenni_zmena)[:2]}% a měsíční změna {str(byty_mesicni_zmena)[:2]}%.
    
    Více dat k dispozici v tabulce: https://1drv.ms/x/s!Alhi9q_byRn2hIdf335IerpOOs3LrA?e=HksUFS
    nebo word dokumentu: https://1drv.ms/w/s!Alhi9q_byRn2hIdtl80bcdF4Bzpr7w?e=YAErft
    další statistiky(ceny, doba inzerce, počet nových inzercí, navšťevy za den ...) na https://www.sreality.cz/ceny-nemovitosti
    """
    return email_text
    
def send_mail(iteration, email_text):
    email_user =  'jan.halik.dev@gmail.com'
    email_password = 'zdomxaomlnpshxvs'
    email_reciever = 'halik.honza2@gmail.com'
    
    subject = "sreality report"
    body = email_text
    
    em = EmailMessage()
    em["From"] =  email_user
    em["To"] = email_reciever
    em["subject"] = subject
    em.set_content(body)
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context = context) as smtp:
        smtp.login(email_user, email_password)
        if not test_mode:
            smtp.sendmail(email_user, email_reciever, em.as_string())
    print("email sent")
    
def main():
    byty, domy = chrome(urls)
    byty, domy, iteration = excel(byty, domy, file_path)
    email_text = statistiky(iteration)
    send_mail(iteration,email_text)
    print("DONE")

if __name__ == "__main__" :
    main()
