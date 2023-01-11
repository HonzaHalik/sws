import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import element_to_be_clickable
import openpyxl
from datetime import date
import re


#TODO
    #reporty typu tento tyden bylo bytu v nabidce x, o x procent mene nez minuly tyden, o x procent mene nez pred mesicem
    #rozesilani techto reportu, excel tabulky
    #grafy v excel tabulce


file_path = fr"c:\Users\halik\OneDrive\Plocha\remax web scraper\remax.xlsx"
url_byty = "https://www.sreality.cz/hledani/byty"
url_domy = "https://www.sreality.cz/hledani/domy"
urls = ["https://www.sreality.cz/hledani/byty", "https://www.sreality.cz/hledani/domy"]
chars = fr"abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ!$%&\'()*+,-./:;<=>?@[\\]^_`~ \t\n\r\x0b\x0c"
sleep = 1


def chrome(urls):
    for url in urls:
        # Set up the webdriver
        driver = webdriver.Chrome()
        # create a new Chrome browser instance in headless mode
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        driver = webdriver.Chrome(chrome_options=options)
        # Navigate to the URL
        driver.get(url)
        # Wait for the page to load
        time.sleep(sleep)
        # Find the button with the specified CSS class and click it
        button = driver.find_element(By.CSS_SELECTOR, ".praha .tspan")
        button.click()
        # Wait for the page to load
        time.sleep(sleep)
        # Get all the text on the page
        page_text = driver.find_element(By.TAG_NAME, "body").text
        # Close the webdriver
        driver.close()
        if url == url_byty:
            index = page_text.find("inzerátů")
            byty = page_text[index - 7 : index]
            byty = re.sub(r'[^0-9]', '', byty)
            print(byty)
            byty = int(byty.replace(" ", ""))
            
            
        elif url == url_domy:
            index = page_text.find("inzerátů")
            domy = page_text[index - 7 : index]
            domy = re.sub(r'[^0-9]', '', domy)
            print(domy)
            domy = int(domy.replace(" ", ""))
    return byty,domy

def excel(byty, domy, file_path):
    # Open the workbook and specify the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # Specify the cell to read the value from
    cell = sheet['A1']
    # Read the value of the cell and ad one to it
    iteration = cell.value
    iteration+=1
    cell.value = iteration
    byty_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+2)
    domy_cell = sheet.cell(row=cell.row+iteration, column=cell.col_idx+3)
    date_cell = sheet.cell(cell.row+iteration, cell.col_idx+1)
    # Write the value num to the new cell
    byty_cell.value = byty
    domy_cell.value = domy
    date_cell.value = date.today()
    # Save the changes to the workbook
    wb.save(file_path)
    print(f"file saved to {file_path}")
    
def main():
    byty, domy = chrome(urls)
    excel(byty, domy, file_path)
    print("DONE")

main()
